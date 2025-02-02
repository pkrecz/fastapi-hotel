from datetime import datetime, UTC
from fastapi import UploadFile, HTTPException, status
from openpyxl import load_workbook
from pathlib import Path
from config.settings import settings


class NoDataInFileException(Exception):
    pass


async def convert_xlsx_to_list_of_dict(file: UploadFile) -> list:

    try:
        list_of_dict = list()
        workbook = load_workbook(
                                    filename=file.file,
                                    read_only=True)
        sheet = workbook.active
        if sheet.max_row <= 1:
            raise NoDataInFileException
        for i in range(2, sheet.max_row + 1):
            row = dict()
            for j in range(1, sheet.max_column + 1):
                column_name = (sheet.cell(row=1, column=j).value).lower()
                row_data = sheet.cell(row=i, column=j).value
                row.update({column_name: row_data})
            list_of_dict.append(row)
        workbook.close()
        return list_of_dict
    except NoDataInFileException:
        raise HTTPException(
                                detail="No data in file - empty or only headers.",
                                status_code=status.HTTP_404_NOT_FOUND)
    except:
        raise HTTPException(
                                detail="File conversion error.",
                                status_code=status.HTTP_400_BAD_REQUEST)


async def save_log_file(prefix: str, line_context: str):
    current_date = datetime.now(tz=UTC).strftime("%Y_%m_%d")
    file_name = f"{prefix}_{current_date}.log"
    full_path_file = Path(settings.LOG_FILES, file_name)
    file = open(file=full_path_file, mode="a")
    file.write(line_context)
    file.close()
